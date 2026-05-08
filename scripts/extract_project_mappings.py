"""
extract_project_mappings.py
---------------------------
Extract missing Oracle→Jira project mappings from validation output.

This script analyzes the validation results to identify:
1. All unique projects in the timesheet (Oracle names)
2. What Jira projects they should map to (from Jira Tickets sheet)
3. Which mappings are missing in Project Edits sheet
4. Generates CSV/Excel with suggested mappings

Usage:
    python scripts/extract_project_mappings.py \
      --validation-output "correction_output.xlsx" \
      --tickets-file "Tickets and People.xlsx" \
      --output "missing_mappings.xlsx"
"""

import argparse
import sys
from pathlib import Path
from collections import defaultdict

import pandas as pd
import openpyxl


def extract_oracle_projects_from_validation(validation_file: str) -> dict:
    """
    Read validation output (All Entries sheet).
    Return dict: {oracle_project_code: set(jira_projects_referenced)}
    """
    print(f"📖 Reading validation output: {validation_file}")
    df = pd.read_excel(validation_file, sheet_name="All Entries")

    projects = defaultdict(set)

    for idx, row in df.iterrows():
        oracle_proj = str(row.get("Project #", "")).strip()
        jira_proj = str(row.get("Jira Oracle Project", "")).strip()

        if oracle_proj and jira_proj:
            projects[oracle_proj].add(jira_proj)

    print(f"   Found {len(projects)} unique Oracle projects")
    return dict(projects)


def load_jira_tickets_mapping(tickets_file: str) -> dict:
    """
    Load Jira Tickets sheet.
    Return dict: {oracle_project_name: jira_project_name}
    """
    print(f"📖 Reading Jira Tickets: {tickets_file}")
    wb = openpyxl.load_workbook(tickets_file, read_only=True, data_only=True)

    if "Tickets" not in wb.sheetnames:
        print(f"   ERROR: 'Tickets' sheet not found in {tickets_file}")
        return {}

    ws = wb["Tickets"]
    mappings = {}

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:  # Skip header
            continue

        # Columns: [0]=Ticket, [2]=Oracle Project, [3]=Jira Project
        oracle_proj = str(row[2]).strip() if len(row) > 2 and row[2] else None
        jira_proj = str(row[3]).strip() if len(row) > 3 and row[3] else None

        if oracle_proj and jira_proj:
            mappings[oracle_proj] = jira_proj

    print(f"   Found {len(mappings)} Jira→Oracle mappings in Tickets sheet")
    return mappings


def load_existing_mappings(project_edits_file: str) -> dict:
    """
    Load existing Project Edits sheet.
    Return dict: {oracle_name: jira_name}
    """
    print(f"📖 Reading existing Project Edits: {project_edits_file}")
    wb = openpyxl.load_workbook(project_edits_file, read_only=True, data_only=True)

    if "Project Edits" not in wb.sheetnames:
        print(f"   WARNING: 'Project Edits' sheet not found")
        return {}

    ws = wb["Project Edits"]
    mappings = {}

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:  # Skip header
            continue

        oracle_name = str(row[0]).strip() if len(row) > 0 and row[0] else None
        jira_name = str(row[1]).strip() if len(row) > 1 and row[1] else None

        if oracle_name and jira_name:
            mappings[oracle_name.upper()] = jira_name

    print(f"   Found {len(mappings)} existing mappings")
    return mappings


def analyze_missing_mappings(validation_file: str, tickets_file: str,
                            existing_mappings: dict) -> tuple:
    """
    Analyze and find missing mappings.
    Returns (missing_list, unmatched_list)
    """
    oracle_projects = extract_oracle_projects_from_validation(validation_file)
    jira_mappings = load_jira_tickets_mapping(tickets_file)

    missing = []
    unmatched = []

    for oracle_code in sorted(oracle_projects.keys()):
        # Check if this code already has a mapping
        if oracle_code.upper() in existing_mappings:
            print(f"   ✓ {oracle_code}: already mapped")
            continue

        # Get all Jira projects referenced for this Oracle code
        jira_projects = oracle_projects[oracle_code]

        if len(jira_projects) == 1:
            jira_name = list(jira_projects)[0]
            missing.append({
                "oracle_project_code": oracle_code,
                "suggested_jira_project": jira_name,
                "status": "suggested",
                "reason": "From validation tickets"
            })
            print(f"   ⚠ {oracle_code} → {jira_name} (suggested)")
        elif len(jira_projects) > 1:
            unmatched.append({
                "oracle_project_code": oracle_code,
                "jira_projects_found": " | ".join(sorted(jira_projects)),
                "status": "ambiguous",
                "reason": f"Multiple Jira projects: {len(jira_projects)}"
            })
            print(f"   ❌ {oracle_code}: AMBIGUOUS ({len(jira_projects)} Jira projects)")
        else:
            unmatched.append({
                "oracle_project_code": oracle_code,
                "jira_projects_found": "NONE",
                "status": "unmapped",
                "reason": "No Jira ticket found for this project"
            })
            print(f"   ⚠ {oracle_code}: NO JIRA MAPPING FOUND")

    return missing, unmatched


def main():
    parser = argparse.ArgumentParser(
        description="Extract missing Oracle→Jira project mappings"
    )
    parser.add_argument("--validation-output", required=True,
                       help="Path to correction_output.xlsx from validation")
    parser.add_argument("--tickets-file", required=True,
                       help="Path to MS Weekly Hrs workbook (for Tickets sheet)")
    parser.add_argument("--project-edits-file",
                       help="Path to existing Project Edits file (optional)")
    parser.add_argument("--output", default="project_mappings_suggested.xlsx",
                       help="Output file path")

    args = parser.parse_args()

    # Validate inputs
    for f in [args.validation_output, args.tickets_file]:
        if not Path(f).exists():
            print(f"❌ File not found: {f}")
            sys.exit(1)

    # Load existing mappings if provided
    existing = {}
    if args.project_edits_file and Path(args.project_edits_file).exists():
        existing = load_existing_mappings(args.project_edits_file)

    print("\n" + "="*60)
    print("Analyzing project mappings...")
    print("="*60 + "\n")

    # Analyze
    missing, unmatched = analyze_missing_mappings(
        args.validation_output,
        args.tickets_file,
        existing
    )

    # Generate output
    print(f"\n📊 Summary:")
    print(f"   Missing mappings: {len(missing)}")
    print(f"   Ambiguous mappings: {len(unmatched)}")

    # Write to Excel
    with pd.ExcelWriter(args.output, engine='openpyxl') as writer:
        if missing:
            df_missing = pd.DataFrame(missing)
            df_missing.to_excel(writer, sheet_name='Missing Mappings', index=False)
            print(f"\n✅ Sheet 'Missing Mappings': {len(missing)} rows")

        if unmatched:
            df_unmatched = pd.DataFrame(unmatched)
            df_unmatched.to_excel(writer, sheet_name='Ambiguous', index=False)
            print(f"⚠️  Sheet 'Ambiguous': {len(unmatched)} rows (review manually)")

    print(f"\n📁 Output: {args.output}")
    print("\n💡 Next steps:")
    print("   1. Review the output file")
    print("   2. Copy 'Missing Mappings' to your 'Project Edits' sheet")
    print("   3. Manually resolve 'Ambiguous' entries")
    print("   4. Re-run validation\n")


if __name__ == "__main__":
    main()
