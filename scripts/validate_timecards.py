"""
validate_timecards.py
---------------------
Step 1 — Standalone timecard validation script.

Reads:
  1. Fusion Timecard xlsx  (Sheet1 — 34 columns)
  2. MS Weekly Hrs xlsx    (Tickets sheet required; People, Project Edits optional)

Replicates all Excel VLOOKUP + formula logic in Python, then runs
7 validation rules to auto-detect correction notes.

Outputs:
  correction_output.xlsx  — same format as weekly correction sheets
                            Column A = auto-detected correction note
"""

import re
import sys
import argparse
from datetime import datetime
from pathlib import Path

import openpyxl
import pandas as pd
from rapidfuzz import process as fuzz_process, fuzz

# ---------------------------------------------------------------------------
# Column indices (0-based) in Fusion Sheet1
# Updated: May 2026 — Fee Arrangement Type added at column R (17)
# This shifted all columns after it down by 1
# ---------------------------------------------------------------------------
F_PERIOD    = 0   # Timecard Period
F_EMP_NAME  = 1   # Employee Name
F_STATUS    = 2   # Timecard Status
F_DATE      = 3   # Date
F_EMP_NUM   = 4   # Employee Number
F_EMP_BU    = 5   # Employee Business Unit
F_FT_PT     = 6   # Full Time / Part Time
F_PTYPE     = 7   # User Person Type
F_JOB       = 8   # Person Job
F_EXEMPT    = 9   # Exempt / Non Exempt
F_DEPT      = 10  # Department
F_MGR       = 11  # Employee Manager Name
F_TA_MGR    = 12  # Time and Absence Manager Name
F_PM        = 13  # Project Manager Name
F_PROJ_BU   = 14  # Project Business Unit
F_PROJ_NUM  = 15  # Project Number
F_PROJ_NAME = 16  # Project Name (Customer/Job)
F_FEE_ARR   = 17  # Fee Arrangement Type (NEW)
F_TASK_NUM  = 18  # Task Number (was 17)
F_TASK_NAME = 19  # Task Name (was 18)
F_BILLABLE  = 20  # Billable Task (was 19)
F_PROJ_TYPE = 21  # Project Type (was 20)
F_PAY_TYPE  = 22  # Payroll Time Type (was 21)
# Columns 23-25: State, County, City (new)
# Column 26: Absence (new)
F_HOURS     = 27  # Hours Worked (was 26)
F_ABS_HRS   = 28  # Absence Hours (was 27)
F_TOTAL_HRS = 29  # Total Hours (was 28)
F_MEMO      = 30  # Comments (by line) (was 29) ← TICKET EXTRACTION

# ---------------------------------------------------------------------------
# Ticket extraction — replicates the Excel formula exactly
# Formula logic:
#   IF memo starts with digit  → extract number before first dash
#   ELSE                       → try extract before 2nd dash (e.g. AHMS-8 - desc)
#                                fallback: before 1st dash
# ---------------------------------------------------------------------------
TICKET_PATTERN = re.compile(r'\b([A-Z][A-Z0-9]+-\d+)\b')
SPACES_IN_TICKET = re.compile(r'\b([A-Z][A-Z0-9]+)\s+-\s*(\d+)\b')
EM_DASH = '\u2013'   # –
EM_DASH2 = '\u2014'  # —


def extract_ticket(memo: str):
    """
    Returns (ticket, raw_extracted, issues) where issues is a list of
    format problems found before a clean ticket could be extracted.
    """
    if not memo or not str(memo).strip():
        return None, None, []

    memo = str(memo).strip()
    issues = []
    # Convert em dashes to regular hyphens for ticket extraction
    # (we'll check the original memo to see if separator uses em dash)
    memo_clean = memo.replace(EM_DASH, '-').replace(EM_DASH2, '-')

    # Check for spaces inside ticket: "CLSMS - 7" or "CLSMS -7"
    space_match = SPACES_IN_TICKET.search(memo_clean)
    if space_match:
        issues.append('spaces_in_ticket')
        # Fix spaces for extraction purposes
        fixed = space_match.group(1) + '-' + space_match.group(2)
        memo_clean = memo_clean[:space_match.start()] + fixed + memo_clean[space_match.end():]

    # Find all ticket-like patterns
    all_tickets = TICKET_PATTERN.findall(memo_clean)

    if not all_tickets:
        return None, None, issues

    if len(all_tickets) > 1:
        issues.append('multiple_tickets')

    ticket = all_tickets[0]

    # Check separator after ticket: should be " - " not "(" or nothing
    # Find position of ticket in original memo to check for em dash in separator
    pos = memo.find(ticket)
    if pos >= 0:
        after_original = memo[pos + len(ticket):]
        # Check if separator uses em dash instead of regular hyphen
        if after_original and after_original[0] in (EM_DASH, EM_DASH2):
            issues.append('em_dash')

    # Use memo_clean for remaining checks
    pos = memo_clean.find(ticket)
    after = memo_clean[pos + len(ticket):]

    after_stripped = after.lstrip()
    if after_stripped and not after_stripped.startswith('-') and not after_stripped.startswith('('):
        # Has description but no proper separator
        if len(after_stripped) > 1:
            issues.append('no_dash_separator')
    elif after_stripped.startswith('('):
        issues.append('no_dash_separator')

    return ticket, memo, issues


# ---------------------------------------------------------------------------
# Load lookup tables from MS Weekly Hrs file
# ---------------------------------------------------------------------------

def load_jira_lookups(jira_path: str):
    wb = openpyxl.load_workbook(jira_path, read_only=True, data_only=True)

    # --- Tickets sheet (required) ---
    ws_tickets = wb['Tickets']
    tickets = {}

    # Build header map to find columns dynamically
    headers = None
    col_map = {}

    for i, row in enumerate(ws_tickets.iter_rows(values_only=True)):
        # Parse header row (first row)
        if i == 0:
            headers = [str(h).strip().lower() if h else '' for h in row]
            # Find key columns dynamically
            for col_idx, header in enumerate(headers):
                if 'key' in header:
                    col_map['key'] = col_idx
                elif 'issue type' in header:
                    col_map['issue_type'] = col_idx
                elif 'oracle' in header and 'project' in header:
                    col_map['oracle_project'] = col_idx
                elif 'jira' in header and 'project' in header:
                    col_map['jira_project'] = col_idx
                elif 'label' in header:
                    col_map['labels'] = col_idx
                elif 'parent' in header:
                    col_map['parent'] = col_idx
            continue

        # Skip if we couldn't find key column
        if 'key' not in col_map:
            continue

        # Extract key from mapped column
        key = row[col_map['key']] if col_map['key'] < len(row) else None
        if not key:
            continue

        # Extract other fields using column map
        tickets[str(key).strip()] = {
            'oracle_project': str(row[col_map.get('oracle_project', 99)]).strip() if col_map.get('oracle_project', 99) < len(row) and row[col_map.get('oracle_project', 99)] else '',
            'jira_project':   str(row[col_map.get('jira_project', 99)]).strip() if col_map.get('jira_project', 99) < len(row) and row[col_map.get('jira_project', 99)] else '',
            'labels':         str(row[col_map.get('labels', 99)]).strip() if col_map.get('labels', 99) < len(row) and row[col_map.get('labels', 99)] else '',
            'issue_type':     str(row[col_map.get('issue_type', 99)]).strip() if col_map.get('issue_type', 99) < len(row) and row[col_map.get('issue_type', 99)] else '',
            'parent':         str(row[col_map.get('parent', 99)]).strip() if col_map.get('parent', 99) < len(row) and row[col_map.get('parent', 99)] else '',
        }

    # --- People sheet (optional) ---
    people = {}
    try:
        ws_people = wb['People']
        for i, row in enumerate(ws_people.iter_rows(values_only=True)):
            if i == 0:
                continue
            emp_num = row[0]
            if not emp_num:
                continue
            people[str(emp_num).strip()] = {
                'name':  str(row[1]).strip() if len(row) > 1 and row[1] else '',
                'email': str(row[2]).strip() if len(row) > 2 and row[2] else '',
            }
    except KeyError:
        print("People sheet not found — emails will be unavailable")
    except IndexError:
        print("People sheet has insufficient columns — some data may be missing")

    # --- Project Edits sheet (optional, Oracle name → Jira name mapping) ---
    project_mapping = {}
    try:
        ws_proj = wb['Project Edits']
        for i, row in enumerate(ws_proj.iter_rows(values_only=True)):
            if i == 0:
                continue
            oracle_name = row[0] if len(row) > 0 else None
            jira_name   = row[1] if len(row) > 1 else None
            if oracle_name and jira_name:
                project_mapping[str(oracle_name).strip()] = str(jira_name).strip()
    except KeyError:
        print("Project Edits sheet not found — project mapping will be unavailable")
    except IndexError:
        print("Project Edits sheet has insufficient columns — project mapping may be incomplete")

    wb.close()
    print(f"Loaded: {len(tickets)} Jira tickets | {len(people)} employees | {len(project_mapping)} project mappings")
    return tickets, people, project_mapping


def load_jira_lookups_from_rds():
    """
    Load Jira tickets and people from AWS RDS PostgreSQL.
    Returns (tickets, people, project_mapping) in same format as load_jira_lookups().
    """
    try:
        from app.db.connection import DatabaseManager
        from app.config import settings
        from sqlalchemy import text

        # Initialize DB connection
        DatabaseManager.init(db_url=settings.db_url)

        with DatabaseManager.get_engine().connect() as conn:
            # Load Jira tickets
            tickets = {}
            result = conn.execute(text('SELECT key, oracle_project_name FROM jira_tickets'))
            for row in result:
                key = str(row[0]).strip()
                oracle_proj = str(row[1]).strip() if row[1] else ''
                tickets[key] = {
                    'oracle_project': oracle_proj,
                    'jira_project':   '',
                    'labels':         '',
                    'issue_type':     '',
                    'parent':         '',
                }

            # Load people (employees)
            people = {}
            result = conn.execute(text('SELECT employee_number, employee_name, email_address FROM people'))
            for row in result:
                emp_num = str(row[0]).strip()
                emp_name = str(row[1]).strip() if row[1] else ''
                email = str(row[2]).strip() if row[2] else ''
                people[emp_num] = {
                    'name':  emp_name,
                    'email': email,
                }

            # Project mapping (empty for now — can be enhanced later)
            project_mapping = {}

        DatabaseManager.close()
        print(f"Loaded from RDS: {len(tickets)} Jira tickets | {len(people)} employees")
        return tickets, people, project_mapping

    except Exception as e:
        print(f"Error loading from RDS: {e}")
        print("Falling back to Excel file loading...")
        return None, None, None


# ---------------------------------------------------------------------------
# Filter criteria — replicates Alison's manual Excel filter
# From transcript: filter by PM name + Department (EA-OR), skip Canada/EPM/DBA
# ---------------------------------------------------------------------------

# PMs included for ALL their rows (from BIP report filter screenshot)
AUDIT_PROJECT_MANAGERS = {
    'Meine, Laura',
    'Monahan, Maureen',
    'Tounkara, Youssouf',
    'Barker, Sherrie',
    'Cox, Alison',
}

# Dhiraj is included ONLY for this specific department
DHIRAJ_PM_NAME  = 'Gadia, Dhiraj'
DHIRAJ_DEPT     = 'EA-OR Managed Services OFA'

# Special project always included regardless of PM
SPECIAL_PROJECTS = {
    'OPUS HOLDING LLC-Oracle EPM Support-1003.0',
}

# Employees to exclude from results (even if their PM is in scope)
EXCLUDE_EMPLOYEES = ('Alison', 'Bharath')

# Shared/internal project prefixes — skip project comparison for these
SHARED_PROJECT_PREFIXES = ('SHNBADM', 'OFAINT', 'GOLD')

# Projects where audit does not apply (Yusuf handles separately per transcript)
SKIP_PROJECT_PREFIXES = ('PPS',)


def should_include_row(row, pm_filter: set = None) -> bool:
    """
    Replicates the exact Fusion BIP report filter (from filter screenshot):
      - Include if PM in {Laura, Sherrie, Youssouf, Maureen, Alison}
      - Include if PM = Dhiraj AND dept = EA-OR Managed Services OFA
      - Include if project name = OPUS HOLDING LLC-Oracle EPM Support-1003.0
      - Exclude if employee name contains Alison or Bharath
    pm_filter: UI override — if provided, replaces AUDIT_PROJECT_MANAGERS.
    """
    allowed_pms  = pm_filter if pm_filter is not None else AUDIT_PROJECT_MANAGERS
    pm           = str(row[F_PM]       or '').strip()
    dept         = str(row[F_DEPT]     or '').strip()
    proj_name    = str(row[F_PROJ_NAME] or '').strip()
    emp_name     = str(row[F_EMP_NAME]  or '').strip()

    # Always exclude specific employees
    for excl in EXCLUDE_EMPLOYEES:
        if excl.lower() in emp_name.lower():
            return False

    # PM in main audit list
    if pm in allowed_pms:
        return True

    # Dhiraj only for EA-OR Managed Services OFA
    if pm == DHIRAJ_PM_NAME and dept == DHIRAJ_DEPT:
        return True

    # Special project always included
    if proj_name in SPECIAL_PROJECTS:
        return True

    return False


def is_shared_project(proj_num: str) -> bool:
    """Shared/internal projects — skip project comparison check."""
    if not proj_num:
        return False
    return any(proj_num.startswith(p) for p in SHARED_PROJECT_PREFIXES)


def is_skipped_project(proj_num: str) -> bool:
    """Projects Yusuf handles — skip entirely."""
    if not proj_num:
        return False
    return any(proj_num.startswith(p) for p in SKIP_PROJECT_PREFIXES)


# ---------------------------------------------------------------------------
# Load Fusion timecard rows
# ---------------------------------------------------------------------------

def load_fusion(fusion_path: str, pm_filter: set = None):
    wb = openpyxl.load_workbook(fusion_path, read_only=True, data_only=True)
    ws = wb['Sheet1']
    rows = []
    skipped_filter = 0
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        if should_include_row(row, pm_filter):
            rows.append(row)
        else:
            skipped_filter += 1
    wb.close()
    print(f"Loaded: {len(rows)} Fusion rows (filtered in) | {skipped_filter} excluded (other depts/PMs)")
    return rows


# ---------------------------------------------------------------------------
# Fuzzy match — find closest Jira ticket key
# Used for: "Should be GSHMS-499" when GSHSM-499 was entered
# ---------------------------------------------------------------------------

def fuzzy_find_ticket(extracted: str, ticket_keys: list, threshold=80):
    if not extracted or not ticket_keys:
        return None
    result = fuzz_process.extractOne(
        extracted,
        ticket_keys,
        scorer=fuzz.ratio,
        score_cutoff=threshold
    )
    if result:
        return result[0]
    return None


# ---------------------------------------------------------------------------
# Core validation — one row at a time
# Returns (correction_note, detail)
# ---------------------------------------------------------------------------

def validate_row(fusion_row, tickets, people, project_mapping, ticket_keys_list):
    emp_num   = str(fusion_row[F_EMP_NUM]).strip() if fusion_row[F_EMP_NUM] else ''
    emp_name  = str(fusion_row[F_EMP_NAME]).strip() if fusion_row[F_EMP_NAME] else ''
    status    = str(fusion_row[F_STATUS]).strip()   if fusion_row[F_STATUS]   else ''
    proj_num  = str(fusion_row[F_PROJ_NUM]).strip() if fusion_row[F_PROJ_NUM] else ''
    proj_name = str(fusion_row[F_PROJ_NAME]).strip()if fusion_row[F_PROJ_NAME]else ''
    task_name = str(fusion_row[F_TASK_NAME]).strip()if fusion_row[F_TASK_NAME]else ''
    date      = fusion_row[F_DATE]
    hours     = fusion_row[F_HOURS] or 0
    memo      = str(fusion_row[F_MEMO]).strip() if fusion_row[F_MEMO] else ''

    # Skip projects Yusuf handles (PPS etc.)
    if is_skipped_project(proj_num):
        return None, '', None, None, None

    # R1 — No memo
    # For shared/internal projects (SHNBADM, GOLD) blank memo is still flagged
    if not memo:
        return 'No memo', '', None, None, None

    # Extract ticket + format issues
    ticket, raw_memo, fmt_issues = extract_ticket(memo)

    # R2 — Has text but no ticket pattern found
    if ticket is None and not fmt_issues:
        return 'Need ticket #', '', None, None, None

    # R3 — Spaces inside ticket number
    if 'spaces_in_ticket' in fmt_issues:
        return 'Remove spaces in ticket #', f'e.g. "{memo[:30]}"', ticket, None, None

    # R4 — Em dash instead of hyphen
    if 'em_dash' in fmt_issues:
        return 'Edit long dash to a short dash', '', ticket, None, None

    # R7 — Multiple tickets (skip for shared projects like SHNBADM)
    if 'multiple_tickets' in fmt_issues:
        if not is_shared_project(proj_num):
            return 'Use only one ticket per entry', '', ticket, None, None

    # R4b — No dash separator after ticket number
    if 'no_dash_separator' in fmt_issues:
        return 'Add dash after ticket #', f'e.g. "{ticket} - description"', ticket, None, None

    # If we still have no ticket after all checks
    if not ticket:
        return 'Need ticket #', '', None, None, None

    # --- Ticket found — now look up in Jira ---
    jira_info = tickets.get(ticket)

    # R5 — Ticket not found in Jira
    if not jira_info:
        # Try fuzzy match
        similar = fuzzy_find_ticket(ticket, ticket_keys_list)
        if similar and similar != ticket:
            return f'Should be {similar}', f'(entered: {ticket})', ticket, similar, None
        return 'Check ticket # - not found in Jira', f'Ticket: {ticket}', ticket, None, None

    # --- Ticket found — check project match ---
    jira_oracle_project = jira_info['oracle_project']

    # Shared/internal projects (SHNBADM, OFAINT, GOLD): skip project comparison
    # Alison explicitly excludes these from BAD review per meeting transcript
    if is_shared_project(proj_num):
        return None, '', ticket, None, jira_info

    # Apply project mapping: Oracle current name → Jira project name
    mapped_proj_name = project_mapping.get(proj_name, proj_name)

    # R6 — Project mismatch (replicates =IF(O2=D2,"GOOD","BAD"))
    if jira_oracle_project and jira_oracle_project != mapped_proj_name:
        # Detect RM vs MS mismatch
        is_rm_ticket = 'RM-' in ticket or 'RM' in ticket.split('-')[0].upper()[-2:]
        if is_rm_ticket and 'Release Management' not in proj_name:
            correction = 'Bill to RM project'
        else:
            correction = f'Ticket is for {_short_project(jira_oracle_project)}'
        return correction, f'Jira: {jira_oracle_project[:60]} | Entry: {proj_name[:40]}', ticket, None, jira_info

    # All checks passed
    return None, '', ticket, None, jira_info


def _short_project(oracle_project_name: str) -> str:
    """Extract short client name from full Oracle project string."""
    if not oracle_project_name:
        return ''
    # e.g. "NORTH MEMORIAL REGIONAL HEALTHCARE SYS-Managed Services-1001.0" → "North Memorial"
    parts = oracle_project_name.split('-')
    client = parts[0].strip().title()
    return client[:40]


def build_auto_project_mapping(fusion_rows: list, tickets: dict) -> dict:
    """
    Pre-scan fusion rows to auto-infer project name mappings.

    When a valid Jira ticket is found in a row, we know:
      Fusion Customer/Job → should match → Jira oracle_project

    If they differ, record the mapping: {fusion_name: jira_name}
    Explicit Project Edits entries will override these in run().

    Returns dict of auto-inferred mappings.
    """
    auto_map = {}
    conflicts = {}

    for row in fusion_rows:
        # Skip rows excluded from validation (same guard as main loop)
        if not row[F_EMP_NUM] or not row[F_DATE]:
            continue

        proj_num = str(row[F_PROJ_NUM] or '').strip()
        if is_skipped_project(proj_num) or is_shared_project(proj_num):
            continue

        # Extract ticket from memo
        memo = str(row[F_MEMO] or '').strip()
        ticket, _, _ = extract_ticket(memo)
        if not ticket or ticket not in tickets:
            continue

        # Get both sides of the comparison
        proj_name = str(row[F_PROJ_NAME] or '').strip()
        jira_oracle = tickets[ticket]['oracle_project']

        # Skip if either is empty or they already match
        if not proj_name or not jira_oracle or proj_name == jira_oracle:
            continue

        # Record mapping
        if proj_name not in auto_map:
            auto_map[proj_name] = jira_oracle
        elif auto_map[proj_name] != jira_oracle:
            # Conflict: same Fusion name maps to multiple Jira projects
            # Mark for removal and skip
            conflicts[proj_name] = True
            if proj_name in auto_map:
                del auto_map[proj_name]

    if conflicts:
        print(f"  Auto-map: {len(conflicts)} ambiguous project(s) skipped")

    return auto_map


# ---------------------------------------------------------------------------
# Main — process all rows and write output
# ---------------------------------------------------------------------------

def run(fusion_path: str, jira_path: str = None, output_path: str = None, pm_filter: set = None, use_rds: bool = False):
    """
    Run validation on timesheet data.

    Args:
        fusion_path: Path to Fusion timecard Excel file
        jira_path: Path to Jira/People Excel file (optional if use_rds=True)
        output_path: Path to write correction output Excel
        pm_filter: Optional set of PM names to include
        use_rds: If True, load Jira data from RDS PostgreSQL instead of Excel
    """
    print(f"\nLoading Fusion: {Path(fusion_path).name}")
    fusion_rows = load_fusion(fusion_path, pm_filter)

    # Load Jira lookups from RDS or Excel
    if use_rds:
        print("Loading Jira lookups from RDS PostgreSQL...")
        tickets, people, project_mapping = load_jira_lookups_from_rds()
        if tickets is None:
            # Fallback to Excel if RDS fails
            if not jira_path:
                raise ValueError(
                    "Could not connect to RDS PostgreSQL and no Jira Excel file provided.\n"
                    "Please upload the 'MS Weekly Hrs / Jira Workbook' Excel file."
                )
            print(f"Falling back to Excel: {Path(jira_path).name}")
            tickets, people, project_mapping = load_jira_lookups(jira_path)
    else:
        if not jira_path:
            raise ValueError("Jira Excel file required")
        print(f"Loading Jira lookups: {Path(jira_path).name}")
        tickets, people, project_mapping = load_jira_lookups(jira_path)

    ticket_keys_list = list(tickets.keys())

    # Skip auto-inferred mappings — they cause false positives when tickets are wrongly extracted
    # Only use explicit Project Edits mappings if provided
    merged_project_mapping = project_mapping
    if project_mapping:
        print(f"  Project mappings: {len(project_mapping)} explicit")

    print(f"\nRunning validation on {len(fusion_rows)} rows...")

    results = []
    error_count = 0
    skipped = 0

    for i, row in enumerate(fusion_rows):
        # Skip rows with no employee or no date
        if not row[F_EMP_NUM] or not row[F_DATE]:
            skipped += 1
            continue

        correction, detail, extracted_ticket, suggested_ticket, jira_info = validate_row(
            row, tickets, people, merged_project_mapping, ticket_keys_list
        )

        # Get employee email from People sheet
        emp_num = str(row[F_EMP_NUM]).strip() if row[F_EMP_NUM] else ''
        emp_info = people.get(emp_num, {})
        email    = emp_info.get('email', '')

        entry_date = row[F_DATE]
        if hasattr(entry_date, 'date'):
            entry_date = entry_date.date()

        record = {
            'Corrections Needed':       correction or '',
            'Status':                   str(row[F_STATUS] or ''),
            'Project #':                str(row[F_PROJ_NUM] or ''),
            'Customer/Job':             str(row[F_PROJ_NAME] or ''),
            'Task Name':                str(row[F_TASK_NAME] or ''),
            'Date':                     entry_date,
            'Employee #':               emp_num,
            'Employee':                 str(row[F_EMP_NAME] or ''),
            'Email':                    email,
            'Actual Time':              row[F_HOURS] or 0,
            'Memo':                     str(row[F_MEMO] or ''),
            'Extracted Ticket':         extracted_ticket or '',
            'Suggested Ticket':         suggested_ticket or '',
            'Jira Oracle Project':      jira_info['oracle_project'] if jira_info else '',
            'Project Match':            'GOOD' if (correction is None) else ('BAD' if jira_info else 'N/A'),
            'Error Detail':             detail,
            'Issue Type':               jira_info['issue_type'] if jira_info else '',
            'Labels':                   jira_info['labels'] if jira_info else '',
            'Period':                   str(row[F_PERIOD] or ''),
        }

        if correction:
            error_count += 1

        results.append(record)

    df = pd.DataFrame(results)
    total = len(df)
    # Filter for rows with non-empty correction notes (handles '', None, NaN, whitespace)
    errors_df = df[
        (df['Corrections Needed'].notna()) &
        (df['Corrections Needed'].astype(str).str.strip() != '')
    ].copy()

    print(f"  Total rows processed : {total}")
    print(f"  Rows skipped (blank) : {skipped}")
    print(f"  Rows with errors     : {len(errors_df)}")
    print(f"  Clean rows           : {total - len(errors_df)}")

    # --- Error breakdown ---
    print("\nError breakdown:")
    if not errors_df.empty:
        for err, cnt in errors_df['Corrections Needed'].value_counts().items():
            print(f"  {cnt:4d}  {err}")

    # --- Write output Excel ---
    print(f"\nWriting output: {output_path}")
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:

        # Sheet 1: Errors only (matches weekly correction sheet format)
        errors_df.to_excel(writer, sheet_name='Corrections Needed', index=False)

        # Sheet 2: All rows
        df.to_excel(writer, sheet_name='All Entries', index=False)

        # Sheet 3: Summary
        summary_data = {
            'Category': list(errors_df['Corrections Needed'].value_counts().index),
            'Count':    list(errors_df['Corrections Needed'].value_counts().values),
        }
        pd.DataFrame(summary_data).to_excel(writer, sheet_name='Summary', index=False)

        # Apply formatting on Corrections Needed sheet
        wb_out = writer.book
        ws_out = wb_out['Corrections Needed']

        # Color coding by error category
        from openpyxl.styles import PatternFill, Font

        red    = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
        orange = PatternFill(start_color='FFE4B5', end_color='FFE4B5', fill_type='solid')
        yellow = PatternFill(start_color='FFFACD', end_color='FFFACD', fill_type='solid')
        header_fill = PatternFill(start_color='C74634', end_color='C74634', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)

        # Color header row
        for cell in ws_out[1]:
            cell.fill   = header_fill
            cell.font   = header_font

        red_errors    = {'No memo', 'Need ticket #', 'Check ticket # - not found in Jira'}
        orange_errors = {'Remove spaces in ticket #', 'Add dash after ticket #',
                         'Edit long dash to a short dash', 'Use only one ticket per entry'}

        for row_cells in ws_out.iter_rows(min_row=2):
            note = row_cells[0].value or ''
            if note in red_errors or note.startswith('Check ticket'):
                fill = red
            elif note in orange_errors:
                fill = orange
            elif note.startswith('Should be') or note.startswith('Ticket is for') or note.startswith('Bill to'):
                fill = yellow
            else:
                continue
            for cell in row_cells:
                cell.fill = fill

        # Auto-width columns
        for col in ws_out.columns:
            max_len = max((len(str(c.value or '')) for c in col), default=10)
            ws_out.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    print(f"Done. Output saved to: {output_path}")
    print(f"\nRed   = missing memo or ticket not found")
    print(f"Orange = format issues (spaces, dash)")
    print(f"Yellow = wrong project or wrong ticket #")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Timecard validation script')
    parser.add_argument('--fusion', required=False,
                        default='C:/Users/PraveenBhaskarani/Downloads/Alithya TS/ALITHYA - Timecard Dump by employee Report (50).xlsx',
                        help='Path to Fusion timecard xlsx')
    parser.add_argument('--jira',   required=False,
                        default='C:/Users/PraveenBhaskarani/Downloads/Alithya TS/MS Weekly Hrs_3.14.xlsx',
                        help='Path to MS Weekly Hrs xlsx (Jira lookups)')
    parser.add_argument('--output', required=False,
                        default='C:/Users/PraveenBhaskarani/Downloads/Alithya TS/correction_output.xlsx',
                        help='Output xlsx path')
    args = parser.parse_args()

    run(args.fusion, args.jira, args.output)
