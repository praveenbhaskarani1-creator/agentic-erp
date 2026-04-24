"""
verify_mappings.py
------------------
Verify that Project Edits mappings are loaded correctly.

Usage:
    python scripts/verify_mappings.py \
      --jira-file "Downloads/Tickets and People.xlsx"
"""

import argparse
import sys
from pathlib import Path
import openpyxl


def check_mappings(jira_file: str):
    """Load and verify Project Edits mappings."""

    if not Path(jira_file).exists():
        print(f"[ERROR] File not found: {jira_file}")
        sys.exit(1)

    print(f"[*] Opening: {jira_file}\n")
    wb = openpyxl.load_workbook(jira_file, read_only=True, data_only=True)

    if "Project Edits" not in wb.sheetnames:
        print("[ERROR] 'Project Edits' sheet not found")
        print(f"Available sheets: {wb.sheetnames}")
        sys.exit(1)

    print("[OK] Found 'Project Edits' sheet\n")

    ws = wb["Project Edits"]
    mappings = {}

    print("Reading mappings:")
    print("-" * 70)
    print(f"{'Oracle Code':<20} | {'Jira Project Name':<40}")
    print("-" * 70)

    row_count = 0
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            print(f"Header: {row}\n")
            continue

        oracle_name = row[0]
        jira_name = row[1]

        if not oracle_name or not jira_name:
            continue

        oracle_str = str(oracle_name).strip()
        jira_str = str(jira_name).strip()

        mappings[oracle_str] = jira_str
        row_count += 1

        print(f"{oracle_str:<20} | {jira_str:<40}")

    print("-" * 70)
    print(f"\nTotal mappings loaded: {row_count}\n")

    # Check for specific mappings (use full Customer/Job names, not project codes)
    test_codes = [
        "Childrens Hospital-Managed Services-1004.0",
        "Mount Nittany Medical Center-Managed Services-1004.0",
        "United Regional Healthcare-Managed Services-1004.0"
    ]
    print("Checking specific mappings:")
    print("-" * 70)

    for code in test_codes:
        if code in mappings:
            print(f"[OK] {code:<50} -> {mappings[code]}")
        else:
            print(f"[MISSING] {code:<50} not found in mappings")

    print("-" * 70)

    # Check for case sensitivity issues
    print("\nChecking case variations:")
    print("-" * 70)

    for oracle_code in list(mappings.keys())[:5]:
        variations = [
            oracle_code,
            oracle_code.upper(),
            oracle_code.lower(),
        ]
        print(f"\n  Base: {oracle_code}")
        for var in variations:
            found = var in mappings
            status = "[OK]" if found else "[MISSING]"
            print(f"    {status} {var:<15} -> {mappings.get(var, 'NOT FOUND')}")

    wb.close()


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Verify Project Edits mappings")
    parser.add_argument("--jira-file", required=True,
                       help="Path to Jira workbook with Project Edits sheet")
    args = parser.parse_args()

    check_mappings(args.jira_file)
