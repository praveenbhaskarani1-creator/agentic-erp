"""
scripts/load_ms_lookups.py
--------------------------
Load Jira lookup tables from the MS Weekly Hrs Excel workbook into Oracle ADW
via ORDS REST API (HTTPS port 443 — no wallet needed).

Reads three sheets:
  Tickets       -> ts_jira_tickets    (ticket_key, oracle_project, jira_project, ...)
  People        -> ts_people          (employee_number, employee_name, email)
  Project Edits -> ts_project_mapping (oracle_project_name -> jira_project_name)

Usage:
    python scripts/load_ms_lookups.py --file "C:/path/MS Weekly Hrs_3.14.xlsx"
    python scripts/load_ms_lookups.py --file "..." --dry-run
    python scripts/load_ms_lookups.py --file "..." --preview
    python scripts/load_ms_lookups.py --file "..." --sheets tickets people
"""

import argparse
import logging
import os
import sys
import re
from pathlib import Path

import openpyxl
from dotenv import load_dotenv

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from scripts.oci_db import OrdsDB

load_dotenv()

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger(__name__)

SHARED_PROJECT_PREFIXES = ("SHNBADM", "OFAINT", "GOLD")

TICKETS_COLS = {
    "ticket_key":          0,
    "summary":             1,
    "oracle_project_name": 2,
    "jira_project_name":   3,
    "labels":              4,
    "issue_type":          5,
    "parent":              6,
}
PEOPLE_COLS = {
    "employee_number": 0,
    "employee_name":   1,
    "email":           2,
}
MAPPING_COLS = {
    "oracle_project_name": 0,
    "jira_project_name":   1,
}


def esc(val, max_len=500) -> str:
    """Escape a string value for inline SQL — replace single quotes."""
    if val is None:
        return "NULL"
    s = str(val).strip()[:max_len].replace("'", "''")
    return f"'{s}'"


def read_sheet_rows(wb, sheet_name: str, col_map: dict, skip_header=True):
    if sheet_name not in wb.sheetnames:
        return None, f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"
    ws = wb[sheet_name]
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if skip_header and i == 0:
            continue
        record = {}
        for field, col_idx in col_map.items():
            val = row[col_idx] if col_idx < len(row) else None
            record[field] = str(val).strip() if val is not None else None
        rows.append(record)
    return [r for r in rows if any(v for v in r.values())], None


def log_upload(db: OrdsDB, table_name, filename, attempted, inserted, skipped, status):
    fn = esc(filename[:500])
    tn = esc(table_name)
    db.execute(f"""
        INSERT INTO upload_log
            (table_name, filename, rows_attempted, rows_inserted,
             rows_skipped_duplicate, rows_skipped_filter, status)
        VALUES ({tn}, {fn}, {attempted}, {inserted}, {skipped}, 0, '{status}')
    """)


def load_tickets(wb, filename: str, db: OrdsDB, dry_run: bool) -> dict:
    rows, err = read_sheet_rows(wb, "Tickets", TICKETS_COLS)
    if err:
        log.error(err)
        return {"inserted": 0, "updated": 0, "errors": 0}

    log.info(f"  Tickets sheet: {len(rows)} rows")
    if dry_run:
        log.info(f"  [dry-run] Would upsert {len(rows)} rows into ts_jira_tickets")
        return {"inserted": len(rows), "updated": 0, "errors": 0}

    inserted = updated = errors = 0
    batch = []

    for row in rows:
        if not row.get("ticket_key"):
            continue
        tk  = esc(row.get("ticket_key"), 50)
        sm  = esc(row.get("summary"), 1000)
        op  = esc(row.get("oracle_project_name"), 500)
        jp  = esc(row.get("jira_project_name"), 500)
        lb  = esc(row.get("labels"), 500)
        it  = esc(row.get("issue_type"), 100)
        par = esc(row.get("parent"), 50)
        src = esc(filename, 500)

        sql = f"""MERGE INTO ts_jira_tickets tgt
USING (SELECT {tk} AS ticket_key FROM DUAL) src
ON (UPPER(tgt.ticket_key) = UPPER(src.ticket_key))
WHEN MATCHED THEN UPDATE SET
    summary={sm}, oracle_project_name={op}, jira_project_name={jp},
    labels={lb}, issue_type={it}, parent={par}, source_file={src},
    loaded_at=SYSTIMESTAMP
WHEN NOT MATCHED THEN INSERT
    (ticket_key,summary,oracle_project_name,jira_project_name,labels,issue_type,parent,source_file)
VALUES ({tk},{sm},{op},{jp},{lb},{it},{par},{src})"""
        batch.append(sql)

        # Send in batches of 50
        if len(batch) == 50:
            try:
                results = db.execute_many(batch)
                for r in results:
                    if r == 1:
                        inserted += 1
                    else:
                        updated += 1
            except Exception as e:
                log.error(f"  Batch error: {e}")
                errors += len(batch)
            batch = []

    # Remaining rows
    if batch:
        try:
            results = db.execute_many(batch)
            for r in results:
                if r == 1:
                    inserted += 1
                else:
                    updated += 1
        except Exception as e:
            log.error(f"  Batch error: {e}")
            errors += len(batch)

    log_upload(db, "ts_jira_tickets", filename, inserted + updated + errors,
               inserted, updated, "success" if errors == 0 else "partial")
    return {"inserted": inserted, "updated": updated, "errors": errors}


def load_people(wb, filename: str, db: OrdsDB, dry_run: bool) -> dict:
    rows, err = read_sheet_rows(wb, "People", PEOPLE_COLS)
    if err:
        log.error(err)
        return {"inserted": 0, "updated": 0, "errors": 0}

    log.info(f"  People sheet: {len(rows)} rows")
    if dry_run:
        log.info(f"  [dry-run] Would upsert {len(rows)} rows into ts_people")
        return {"inserted": len(rows), "updated": 0, "errors": 0}

    inserted = updated = errors = 0
    batch = []

    for row in rows:
        if not row.get("employee_number"):
            continue
        en  = esc(row.get("employee_number"), 50)
        nm  = esc(row.get("employee_name"), 200)
        em  = esc(row.get("email"), 200)
        src = esc(filename, 500)

        sql = f"""MERGE INTO ts_people tgt
USING (SELECT {en} AS employee_number FROM DUAL) src
ON (tgt.employee_number = src.employee_number)
WHEN MATCHED THEN UPDATE SET
    employee_name={nm}, email={em}, source_file={src}, loaded_at=SYSTIMESTAMP
WHEN NOT MATCHED THEN INSERT
    (employee_number, employee_name, email, source_file)
VALUES ({en},{nm},{em},{src})"""
        batch.append(sql)

        if len(batch) == 50:
            try:
                results = db.execute_many(batch)
                for r in results:
                    if r == 1: inserted += 1
                    else: updated += 1
            except Exception as e:
                log.error(f"  Batch error: {e}")
                errors += len(batch)
            batch = []

    if batch:
        try:
            results = db.execute_many(batch)
            for r in results:
                if r == 1: inserted += 1
                else: updated += 1
        except Exception as e:
            log.error(f"  Batch error: {e}")
            errors += len(batch)

    log_upload(db, "ts_people", filename, inserted + updated + errors,
               inserted, updated, "success" if errors == 0 else "partial")
    return {"inserted": inserted, "updated": updated, "errors": errors}


def load_project_mapping(wb, filename: str, db: OrdsDB, dry_run: bool) -> dict:
    rows, err = read_sheet_rows(wb, "Project Edits", MAPPING_COLS)
    if err:
        log.error(err)
        return {"inserted": 0, "updated": 0, "errors": 0}

    log.info(f"  Project Edits sheet: {len(rows)} rows")
    if dry_run:
        log.info(f"  [dry-run] Would upsert {len(rows)} rows into ts_project_mapping")
        return {"inserted": len(rows), "updated": 0, "errors": 0}

    inserted = updated = errors = 0
    batch = []

    for row in rows:
        if not row.get("oracle_project_name") or not row.get("jira_project_name"):
            continue
        proj_upper = str(row["oracle_project_name"]).upper()
        is_shared  = "Y" if any(proj_upper.startswith(p) for p in SHARED_PROJECT_PREFIXES) else "N"
        op  = esc(row["oracle_project_name"], 500)
        jp  = esc(row["jira_project_name"], 500)
        src = esc(filename, 500)

        sql = f"""MERGE INTO ts_project_mapping tgt
USING (SELECT UPPER({op}) AS oracle_project_name FROM DUAL) src
ON (UPPER(tgt.oracle_project_name) = src.oracle_project_name)
WHEN MATCHED THEN UPDATE SET
    jira_project_name={jp}, is_shared_project='{is_shared}',
    source_file={src}, loaded_at=SYSTIMESTAMP
WHEN NOT MATCHED THEN INSERT
    (oracle_project_name, jira_project_name, is_shared_project, source_file)
VALUES ({op},{jp},'{is_shared}',{src})"""
        batch.append(sql)

    if batch:
        try:
            results = db.execute_many(batch)
            for r in results:
                if r == 1: inserted += 1
                else: updated += 1
        except Exception as e:
            log.error(f"  Batch error: {e}")
            errors += len(batch)

    log_upload(db, "ts_project_mapping", filename, inserted + updated + errors,
               inserted, updated, "success" if errors == 0 else "partial")
    return {"inserted": inserted, "updated": updated, "errors": errors}


def preview_sheet(wb, sheet_name: str, col_map: dict, n=10):
    rows, err = read_sheet_rows(wb, sheet_name, col_map)
    if err:
        print(f"  ERROR: {err}")
        return
    print(f"\n  [{sheet_name}] -- {len(rows)} data rows. First {min(n, len(rows))}:")
    print("  " + " | ".join(col_map.keys()))
    print("  " + "-" * 80)
    for row in rows[:n]:
        print("  " + " | ".join(str(row.get(k, ""))[:25] for k in col_map))


def main():
    parser = argparse.ArgumentParser(
        description="Load MS Weekly Hrs lookup sheets into Oracle ADW via ORDS"
    )
    parser.add_argument("--file", required=True, help="Path to MS Weekly Hrs XLSX")
    parser.add_argument("--sheets", nargs="+",
        choices=["tickets", "people", "mapping"],
        default=["tickets", "people", "mapping"])
    parser.add_argument("--dry-run",  action="store_true")
    parser.add_argument("--preview",  action="store_true")
    args = parser.parse_args()

    filepath = Path(args.file)
    if not filepath.exists():
        log.error(f"File not found: {filepath}")
        sys.exit(1)

    log.info(f"Opening workbook: {filepath.name}")
    wb = openpyxl.load_workbook(str(filepath), read_only=True, data_only=True)
    log.info(f"  Sheets found: {wb.sheetnames}")

    if args.preview:
        if "tickets" in args.sheets:
            preview_sheet(wb, "Tickets", TICKETS_COLS)
        if "people" in args.sheets:
            preview_sheet(wb, "People", PEOPLE_COLS)
        if "mapping" in args.sheets:
            preview_sheet(wb, "Project Edits", MAPPING_COLS)
        return

    db = OrdsDB()
    log.info(f"DB connected: {db._sql_url}")

    filename = filepath.name
    totals = {}

    if "tickets" in args.sheets:
        log.info("Loading Tickets -> ts_jira_tickets ...")
        totals["tickets"] = load_tickets(wb, filename, db, args.dry_run)

    if "people" in args.sheets:
        log.info("Loading People -> ts_people ...")
        totals["people"] = load_people(wb, filename, db, args.dry_run)

    if "mapping" in args.sheets:
        log.info("Loading Project Edits -> ts_project_mapping ...")
        totals["mapping"] = load_project_mapping(wb, filename, db, args.dry_run)

    print("\n========== Load Result ==========")
    for tbl, result in totals.items():
        ins = result.get("inserted", 0)
        upd = result.get("updated", 0)
        err = result.get("errors", 0)
        print(f"  {tbl:<10}  inserted={ins}  updated={upd}  errors={err}")
    print("=================================")

    if args.dry_run:
        print("\n[dry-run] No rows were inserted.")


if __name__ == "__main__":
    main()
