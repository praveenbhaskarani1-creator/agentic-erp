"""
generate_project_edits.py
------------------------
Extract "Ticket is for X" errors from validation output and
generate Project Edits mappings.

This script:
1. Reads the validation output (Corrections Needed sheet)
2. Finds all "Ticket is for" errors
3. Extracts: Fusion Customer/Job → Jira Oracle Project mappings
4. Creates a new Project Edits sheet in Excel

Usage:
    python scripts/generate_project_edits.py \
      --validation-output "Downloads/correction_output.xlsx" \
      --output "Project_Edits_Generated.xlsx"

    python scripts/generate_project_edits.py \
      --validation-output "Downloads/correction_output.xlsx" \
      --existing-edits "Downloads/Tickets and People.xlsx" \
      --output "Project_Edits_Updated.xlsx"
"""

import argparse
import sys
from pathlib import Path
from collections import OrderedDict

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


def extract_mappings_from_validation(validation_file: str) -> dict:
    """
    Read validation output and extract all "Ticket is for" mappings.

    Returns: {oracle_project_code: jira_project_name, ...}
    """
    print(f"[*] Reading validation output: {validation_file}")

    try:
        df = pd.read_excel(validation_file, sheet_name="Corrections Needed")
    except Exception as e:
        print(f"[ERROR] Error reading sheet: {e}")
        sys.exit(1)

    mappings = OrderedDict()

    print(f"   Found {len(df)} error rows")
    print("\n   Extracting mappings from 'Ticket is for' errors:\n")

    for idx, row in df.iterrows():
        correction = str(row.get("Corrections Needed", "")).strip()

        # Only process "Ticket is for" errors
        if not correction.startswith("Ticket is for"):
            continue

        # Use the full Customer/Job (Fusion project name) and Jira Oracle Project (from ticket)
        oracle_code = str(row.get("Customer/Job", "")).strip()
        jira_project = str(row.get("Jira Oracle Project", "")).strip()

        if oracle_code and jira_project:
            # Store mapping (no uppercasing needed — Customer/Job is already exact)
            key = oracle_code

            if key in mappings:
                # Check for conflicts
                if mappings[key] != jira_project:
                    print(f"   [!] CONFLICT: {oracle_code}")
                    print(f"       Previous: {mappings[key]}")
                    print(f"       Current:  {jira_project}")
                    print(f"       -> Using most recent\n")
                mappings[key] = jira_project
            else:
                mappings[key] = jira_project
                print(f"   [OK] {oracle_code:20s} -> {jira_project}")

    print(f"\n   Total unique mappings: {len(mappings)}\n")
    return mappings


def load_existing_project_edits(workbook_file: str) -> dict:
    """Load existing Project Edits sheet from Excel."""
    if not Path(workbook_file).exists():
        return {}

    print(f"[*] Reading existing Project Edits from: {workbook_file}")

    try:
        wb = openpyxl.load_workbook(workbook_file, read_only=True, data_only=True)

        if "Project Edits" not in wb.sheetnames:
            print("   [!] 'Project Edits' sheet not found\n")
            return {}

        ws = wb["Project Edits"]
        existing = {}

        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:  # Skip header
                continue

            oracle = str(row[0]).strip() if len(row) > 0 and row[0] else None
            jira = str(row[1]).strip() if len(row) > 1 and row[1] else None

            if oracle and jira:
                existing[oracle.upper()] = jira

        print(f"   Found {len(existing)} existing mappings\n")
        return existing

    except Exception as e:
        print(f"   [!] Error reading existing mappings: {e}\n")
        return {}


def merge_mappings(new_mappings: dict, existing_mappings: dict) -> tuple:
    """
    Merge new mappings with existing ones.
    Returns (merged_dict, added_count, updated_count)
    """
    merged = existing_mappings.copy()
    added = 0
    updated = 0

    for oracle_code, jira_project in new_mappings.items():
        if oracle_code not in merged:
            added += 1
            merged[oracle_code] = jira_project
        elif merged[oracle_code] != jira_project:
            updated += 1
            print(f"   Updating: {oracle_code} (was '{merged[oracle_code]}' → now '{jira_project}')")
            merged[oracle_code] = jira_project

    return merged, added, updated


def create_project_edits_excel(mappings: dict, output_file: str):
    """Create Excel file with Project Edits sheet."""
    print(f"\n[*] Creating Project Edits sheet...")

    # Prepare data
    data = {
        "Oracle Name": [code.lower() for code in mappings.keys()],  # Keep original case
        "Jira Name": list(mappings.values())
    }

    df = pd.DataFrame(data)

    # Create Excel writer
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Project Edits', index=False)

        # Format the sheet
        workbook = writer.book
        worksheet = writer.sheets['Project Edits']

        # Header formatting
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Column widths
        worksheet.column_dimensions['A'].width = 25
        worksheet.column_dimensions['B'].width = 40

        # Alternate row colors
        light_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        for i, row in enumerate(worksheet.iter_rows(min_row=2, max_row=len(df) + 1), start=1):
            if i % 2 == 0:
                for cell in row:
                    cell.fill = light_fill

    print(f"   [OK] Created: {output_file}")
    print(f"   Rows: {len(mappings)}")
    return output_file


def main():
    parser = argparse.ArgumentParser(
        description="Generate Project Edits mapping from validation 'Ticket is for' errors"
    )
    parser.add_argument("--validation-output", required=True,
                       help="Path to correction_output.xlsx")
    parser.add_argument("--existing-edits", default=None,
                       help="Path to existing workbook with Project Edits sheet (optional)")
    parser.add_argument("--output", default="Project_Edits_Generated.xlsx",
                       help="Output file path")

    args = parser.parse_args()

    # Validate input
    if not Path(args.validation_output).exists():
        print(f"❌ File not found: {args.validation_output}")
        sys.exit(1)

    print("="*70)
    print("PROJECT EDITS MAPPING GENERATOR")
    print("="*70 + "\n")

    # Extract new mappings from validation errors
    new_mappings = extract_mappings_from_validation(args.validation_output)

    if not new_mappings:
        print("[ERROR] No 'Ticket is for' errors found in validation output")
        sys.exit(1)

    # Load existing mappings if provided
    existing_mappings = {}
    if args.existing_edits:
        existing_mappings = load_existing_project_edits(args.existing_edits)

    # Merge
    if existing_mappings:
        print(f"[*] Merging with existing mappings...\n")
        merged, added, updated = merge_mappings(new_mappings, existing_mappings)
        print(f"   Added:   {added}")
        print(f"   Updated: {updated}")
        print(f"   Total:   {len(merged)}\n")
    else:
        merged = new_mappings

    # Create output
    create_project_edits_excel(merged, args.output)

    print("\n" + "="*70)
    print("NEXT STEPS:")
    print("="*70)
    print(f"""
1. Open the generated file: {args.output}

2. Review the mappings:
   - Oracle Name (left) = Project codes from your timesheet
   - Jira Name (right) = Project names from Jira tickets

3. Copy the data to your MS Weekly Hrs workbook:
   - Open: Downloads/Tickets and People.xlsx
   - Go to: Project Edits sheet
   - Replace with the new mappings

4. Upload to database:
   python scripts/load_ms_lookups.py \\
     --file "Downloads/Tickets and People.xlsx" \\
     --sheets mapping

5. Re-run validation:
   - All "Ticket is for" errors should now be GOOD [OK]
""")


if __name__ == "__main__":
    main()
