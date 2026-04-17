"""
app/tools/export_tool.py
────────────────────────
Exports all 5 query results to a single Excel file.

One sheet per query + a Summary sheet.
User triggers on demand — file downloads via FastAPI endpoint.

Usage:
    from app.tools.export_tool import ExportTool

    tool = ExportTool()
    filepath = tool.export()
    # Returns: "exports/TimeEntry_Validation_Report_2026-03-13.xlsx"
"""

import logging
import os
from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from app.tools.sql_tool import SQLTool

logger = logging.getLogger(__name__)

# ── Styling constants ─────────────────────────────────────────
HEADER_FONT      = Font(name="Arial", bold=True, color="FFFFFF", size=11)
HEADER_FILL      = PatternFill("solid", start_color="1B4F8A")   # navy blue
SUMMARY_FILL     = PatternFill("solid", start_color="0D7377")   # teal
ROW_FILL_ALT     = PatternFill("solid", start_color="F4F7FB")   # light gray
HEADER_ALIGN     = Alignment(horizontal="center", vertical="center")
CENTER_ALIGN     = Alignment(horizontal="center", vertical="center")

# Tab config: query_name → sheet display name
TAB_CONFIG = [
    ("all_entries",  "All Entries"),
    ("blank_memo",   "Blank Memo"),
    ("last_7_days",  "Last 7 Days"),
    ("non_erp_memo", "Non ERP Memo"),
    ("total_count",  "Total Count"),
]

COLUMN_WIDTHS = {
    "id":             8,
    "employee":       25,
    "date":           15,
    "hours":          10,
    "memo":           40,
    "project_number": 18,
    "project_name":   30,
    "total_records":  18,
}


class ExportTool:
    """
    Generates an Excel report from all query results.
    Each query gets its own sheet. Summary sheet shows counts.
    """

    def __init__(self):
        self.sql_tool = SQLTool()

    def export(self, output_dir: str = "exports") -> str:
        """
        Run all queries, write to Excel, return the file path.

        Args:
            output_dir: folder to save the file (created if missing)

        Returns:
            Full file path of the saved .xlsx file
        """
        os.makedirs(output_dir, exist_ok=True)

        timestamp = datetime.now().strftime("%Y-%m-%d")
        filename  = f"TimeEntry_Validation_Report_{timestamp}.xlsx"
        filepath  = os.path.join(output_dir, filename)

        logger.info(f"[export] Starting export → {filepath}")

        # ── Run all 5 queries ─────────────────────────────────
        results = {}
        for query_name, _ in TAB_CONFIG:
            results[query_name] = self.sql_tool.run(query_name)
            logger.info(
                f"[export] {query_name}: "
                f"{results[query_name]['row_count']} rows"
            )

        # ── Build workbook ────────────────────────────────────
        wb = Workbook()
        wb.remove(wb.active)   # remove default empty sheet

        # Summary sheet first
        self._write_summary(wb, results)

        # One sheet per query
        for query_name, sheet_name in TAB_CONFIG:
            self._write_data_sheet(wb, sheet_name, results[query_name])

        wb.save(filepath)
        logger.info(f"[export] Saved → {filepath}")
        return filepath

    def export_to_bytes(self) -> tuple[bytes, str]:
        """
        Export to bytes in memory — used by FastAPI to stream download.

        Returns:
            (bytes_content, filename)
        """
        timestamp = datetime.now().strftime("%Y-%m-%d")
        filename  = f"TimeEntry_Validation_Report_{timestamp}.xlsx"

        results = {}
        for query_name, _ in TAB_CONFIG:
            results[query_name] = self.sql_tool.run(query_name)

        wb = Workbook()
        wb.remove(wb.active)
        self._write_summary(wb, results)
        for query_name, sheet_name in TAB_CONFIG:
            self._write_data_sheet(wb, sheet_name, results[query_name])

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.read(), filename

    # ─────────────────────────────────────────────────────────
    # Sheet writers
    # ─────────────────────────────────────────────────────────

    def _write_summary(self, wb: Workbook, results: dict):
        """Write the Summary sheet — one row per query with counts."""
        ws = wb.create_sheet("Summary")

        # Title row
        ws.merge_cells("A1:D1")
        ws["A1"] = "Agentic Time Entry Validation — Export Report"
        ws["A1"].font      = Font(name="Arial", bold=True, size=13, color="FFFFFF")
        ws["A1"].fill      = SUMMARY_FILL
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        # Generated timestamp
        ws.merge_cells("A2:D2")
        ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws["A2"].font      = Font(name="Arial", italic=True, size=10)
        ws["A2"].alignment = Alignment(horizontal="center")
        ws.row_dimensions[2].height = 18

        # Header row
        headers = ["Query", "Sheet", "Row Count", "Status"]
        for col, header in enumerate(headers, start=1):
            cell            = ws.cell(row=4, column=col, value=header)
            cell.font       = HEADER_FONT
            cell.fill       = HEADER_FILL
            cell.alignment  = HEADER_ALIGN
        ws.row_dimensions[4].height = 20

        # Data rows
        tab_lookup = dict(TAB_CONFIG)
        for row_idx, (query_name, _) in enumerate(TAB_CONFIG, start=5):
            result     = results[query_name]
            sheet_name = tab_lookup[query_name]
            count      = result["row_count"]
            status     = "✓ OK" if result["status"] in ("success", "empty") else "✗ Error"

            ws.cell(row=row_idx, column=1, value=result.get("description", query_name))
            ws.cell(row=row_idx, column=2, value=sheet_name)
            ws.cell(row=row_idx, column=3, value=count)
            ws.cell(row=row_idx, column=4, value=status)

            # Alternate row fill
            if row_idx % 2 == 0:
                for col in range(1, 5):
                    ws.cell(row=row_idx, column=col).fill = ROW_FILL_ALT

            # Center count + status
            ws.cell(row=row_idx, column=3).alignment = CENTER_ALIGN
            ws.cell(row=row_idx, column=4).alignment = CENTER_ALIGN

        # Column widths for summary
        ws.column_dimensions["A"].width = 50
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 12

    def _write_data_sheet(self, wb: Workbook, sheet_name: str, result: dict):
        """Write one data sheet with headers + rows."""
        ws = wb.create_sheet(sheet_name)

        rows = result.get("rows", [])

        # ── No data case ──────────────────────────────────────
        if not rows:
            ws["A1"] = "No records found"
            ws["A1"].font = Font(name="Arial", italic=True, color="888888")
            return

        # ── Headers ───────────────────────────────────────────
        columns = list(rows[0].keys())
        for col_idx, col_name in enumerate(columns, start=1):
            cell           = ws.cell(row=1, column=col_idx, value=col_name.upper())
            cell.font      = HEADER_FONT
            cell.fill      = HEADER_FILL
            cell.alignment = HEADER_ALIGN

        ws.row_dimensions[1].height = 20

        # ── Data rows ─────────────────────────────────────────
        for row_idx, row_data in enumerate(rows, start=2):
            for col_idx, col_name in enumerate(columns, start=1):
                value = row_data.get(col_name)
                cell  = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = Font(name="Arial", size=10)

                # Alternate row shading
                if row_idx % 2 == 0:
                    cell.fill = ROW_FILL_ALT

                # Center id, date, hours columns
                if col_name in ("id", "date", "hours", "total_records"):
                    cell.alignment = CENTER_ALIGN

        # ── Auto column widths ────────────────────────────────
        for col_idx, col_name in enumerate(columns, start=1):
            col_letter = get_column_letter(col_idx)
            width      = COLUMN_WIDTHS.get(col_name, 20)
            ws.column_dimensions[col_letter].width = width

        # ── Freeze header row ─────────────────────────────────
        ws.freeze_panes = "A2"

        # ── Row count footer ──────────────────────────────────
        footer_row = len(rows) + 3
        ws.cell(row=footer_row, column=1,
                value=f"Total rows: {len(rows)}")
        ws.cell(row=footer_row, column=1).font = Font(
            name="Arial", bold=True, italic=True, size=10
        )
