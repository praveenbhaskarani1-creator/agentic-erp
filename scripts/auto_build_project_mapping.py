"""
auto_build_project_mapping.py
-----------------------------
Automatically build Project Edits mapping by analyzing which project codes
actually contain each Jira project's tickets.

Logic:
1. For each ticket in Jira Tickets sheet, note its oracle_project_name
2. For each timesheet entry, extract the ticket code
3. See which project code contains entries for each oracle_project_name
4. Build: oracle_project_name → project_code mapping automatically

This eliminates manual Project Edits maintenance!

Usage:
    python scripts/auto_build_project_mapping.py \
      --fusion-file "Downloads/Time Dump 3-1 to 4-11.xlsx" \
      --jira-file "Downloads/Tickets and People.xlsx" \
      --output "project_mapping_auto.xlsx"
"""

import argparse
import sys
from pathlib import Path
from collections import defaultdict

import pandas as pd
import openpyxl


def load_jira_projects(jira_file: str) -> dict:
    """
    Load Jira Tickets sheet.
    Return: {ticket_key: oracle_project_name}
    """
    print("[*] Loading Jira Tickets...")

    wb = openpyxl.load_workbook(jira_file, read_only=True, data_only=True)

    if "Tickets" not in wb.sheetnames:
        print("[ERROR] Tickets sheet not found")
        sys.exit(1)

    ws = wb["Tickets"]
    tickets = {}

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:  # Skip header
            continue

        ticket_key = str(row[0]).strip() if row[0] else None
        oracle_proj = str(row[2]).strip() if len(row) > 2 and row[2] else None

        if ticket_key and oracle_proj:
            tickets[ticket_key] = oracle_proj

    print(f"   Found {len(tickets)} tickets\n")
    return tickets


def extract_ticket_from_memo(memo: str):
    """Simple ticket extraction (e.g., CHOMS-94 from memo)."""
    import re

    if not memo:
        return None

    memo = str(memo).strip()
    match = re.search(r'\b([A-Z][A-Z0-9]+-\d+)\b', memo)

    if match:
        return match.group(1)
    return None


def analyze_fusion_data(fusion_file: str, jira_tickets: dict) -> dict:
    """
    Analyze Fusion timesheet.
    For each oracle_project_name, find which project codes contain its tickets.

    Return: {oracle_project_name: {project_codes: count}}
    """
    print("[*] Analyzing Fusion timesheet...")

    df = pd.read_excel(fusion_file, sheet_name="Sheet1")

    project_analysis = defaultdict(lambda: defaultdict(int))

    for idx, row in df.iterrows():
        memo = row.get("Comments (by line)", "")
        project_code = str(row.get("Project Number", "")).strip()

        if not memo or not project_code:
            continue

        # Extract ticket from memo
        ticket = extract_ticket_from_memo(memo)

        if not ticket:
            continue

        # Look up ticket in Jira
        if ticket in jira_tickets:
            oracle_proj = jira_tickets[ticket]
            project_analysis[oracle_proj][project_code] += 1

    print(f"   Analyzed {len(df)} rows\n")
    return dict(project_analysis)


def build_mapping(analysis: dict) -> dict:
    """
    From the analysis, build the best mapping.
    For each oracle_project_name, pick the project_code that appears most.

    Return: {oracle_project_name: project_code}
    """
    mapping = {}

    print("[*] Building mapping from analysis:\n")
    print(f"{'Oracle Project':<40} | {'Project Code':<15} | {'Count':<5}")
    print("-" * 70)

    for oracle_proj, project_codes in analysis.items():
        if not project_codes:
            continue

        # Find most common project code
        best_code = max(project_codes.items(), key=lambda x: x[1])
        project_code = best_code[0]
        count = best_code[1]

        mapping[oracle_proj] = project_code

        print(f"{oracle_proj:<40} | {project_code:<15} | {count:<5}")

        # Warn if there are multiple project codes
        if len(project_codes) > 1:
            others = ", ".join(
                f"{code}({cnt})"
                for code, cnt in sorted(project_codes.items(), key=lambda x: -x[1])[1:]
            )
            print(f"   [!] Also found: {others}")

    print("-" * 70)
    print(f"\nTotal mappings: {len(mapping)}\n")

    return mapping


def create_output_excel(mapping: dict, output_file: str):
    """Create Project Edits sheet from mapping."""
    from openpyxl.styles import Font, PatternFill, Alignment

    print(f"[*] Creating {output_file}...")

    data = {
        "Oracle Name": list(mapping.keys()),
        "Jira Name": list(mapping.values())
    }

    df = pd.DataFrame(data)

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Project Edits', index=False)

        ws = writer.sheets['Project Edits']

        # Format
        header_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 15

    print(f"   [OK] Created: {output_file}\n")


def main():
    parser = argparse.ArgumentParser(
        description="Auto-build Project Edits mapping from actual data"
    )
    parser.add_argument("--fusion-file", required=True,
                       help="Fusion timesheet export")
    parser.add_argument("--jira-file", required=True,
                       help="Jira workbook with Tickets sheet")
    parser.add_argument("--output", default="project_mapping_auto.xlsx",
                       help="Output file")

    args = parser.parse_args()

    for f in [args.fusion_file, args.jira_file]:
        if not Path(f).exists():
            print(f"[ERROR] File not found: {f}")
            sys.exit(1)

    print("\n" + "=" * 70)
    print("AUTO-BUILD PROJECT MAPPING")
    print("=" * 70 + "\n")

    # Step 1: Load Jira tickets
    jira_tickets = load_jira_projects(args.jira_file)

    # Step 2: Analyze Fusion data
    analysis = analyze_fusion_data(args.fusion_file, jira_tickets)

    # Step 3: Build mapping
    mapping = build_mapping(analysis)

    if not mapping:
        print("[ERROR] No mappings could be generated")
        sys.exit(1)

    # Step 4: Create output
    create_output_excel(mapping, args.output)

    print("=" * 70)
    print("NEXT STEPS:")
    print("=" * 70)
    print(f"""
1. Review: {args.output}

2. Copy this mapping to your Tickets and People.xlsx:
   - Open: {args.output}
   - Copy Project Edits sheet
   - Open: Tickets and People.xlsx
   - Replace Project Edits sheet
   - Save

3. Re-run validation:
   - All "Ticket is for" errors should now disappear!
""")


if __name__ == "__main__":
    main()
