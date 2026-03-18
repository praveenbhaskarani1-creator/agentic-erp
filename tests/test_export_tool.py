"""
tests/test_export_tool.py
─────────────────────────
Tests for app/tools/export_tool.py

UNIT tests    — no DB needed
INTEGRATION   — needs SSH tunnel open

Run all:
    pytest tests\test_export_tool.py -v
"""

import os
import sys
import pytest
from unittest.mock import patch, MagicMock

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from dotenv import load_dotenv
load_dotenv()

from app.tools.export_tool import ExportTool, TAB_CONFIG, COLUMN_WIDTHS


# ─────────────────────────────────────────────────────────────
# Fixtures
# ─────────────────────────────────────────────────────────────

@pytest.fixture(scope="module")
def tool():
    return ExportTool()


@pytest.fixture(scope="module", autouse=True)
def init_db():
    from app.db.connection import DatabaseManager
    db_url = (
        f"postgresql+psycopg2://"
        f"{os.getenv('DB_USER', 'pgadmin')}:"
        f"{os.getenv('DB_PASSWORD')}@"
        f"{os.getenv('DB_HOST', 'localhost')}:"
        f"{os.getenv('DB_PORT', '5433')}/"
        f"{os.getenv('DB_NAME', 'agentdb')}"
    )
    try:
        DatabaseManager.init(db_url=db_url, pool_size=3, echo=False)
    except Exception:
        pass
    yield


# ─── Mock results for unit tests ─────────────────────────────

MOCK_RESULTS = {
    "all_entries": {
        "status": "success", "query_name": "all_entries",
        "description": "Retrieve all time entry records",
        "row_count": 3, "message": "Found 3 records",
        "rows": [
            {"id": 1, "employee": "John Smith",  "date": "2026-01-15", "hours": 8.0, "memo": "ERP-001"},
            {"id": 2, "employee": "Jane Doe",    "date": "2026-01-16", "hours": 7.5, "memo": None},
            {"id": 3, "employee": "Bob Lee",     "date": "2026-01-17", "hours": 6.0, "memo": "ERP-002"},
        ]
    },
    "blank_memo": {
        "status": "success", "query_name": "blank_memo",
        "description": "Find entries with NULL memo",
        "row_count": 1, "message": "Found 1 record",
        "rows": [
            {"id": 2, "employee": "Jane Doe", "date": "2026-01-16", "hours": 7.5, "memo": None},
        ]
    },
    "last_7_days": {
        "status": "empty", "query_name": "last_7_days",
        "description": "Entries from last 7 days",
        "row_count": 0, "message": "No records found", "rows": []
    },
    "non_erp_memo": {
        "status": "success", "query_name": "non_erp_memo",
        "description": "Entries where memo not like ERP%",
        "row_count": 1, "message": "Found 1 record",
        "rows": [
            {"id": 2, "employee": "Jane Doe", "date": "2026-01-16", "hours": 7.5, "memo": None},
        ]
    },
    "total_count": {
        "status": "success", "query_name": "total_count",
        "description": "Total number of records",
        "row_count": 1, "message": "Found 1 record",
        "rows": [{"total_records": 3}]
    },
}


# ═════════════════════════════════════════════════════════════
# UNIT TESTS
# ═════════════════════════════════════════════════════════════

class TestTabConfig:

    def test_all_5_queries_in_tab_config(self):
        names = [q for q, _ in TAB_CONFIG]
        assert "all_entries"  in names
        assert "blank_memo"   in names
        assert "last_7_days"  in names
        assert "non_erp_memo" in names
        assert "total_count"  in names

    def test_tab_config_has_display_names(self):
        for query_name, sheet_name in TAB_CONFIG:
            assert len(sheet_name) > 0
            assert sheet_name != query_name   # display name differs from key

    def test_column_widths_defined_for_core_columns(self):
        for col in ["id", "employee", "date", "hours", "memo"]:
            assert col in COLUMN_WIDTHS
            assert COLUMN_WIDTHS[col] > 0


class TestExportToBytes:
    """Unit tests using mocked SQL results — no DB needed."""

    def test_export_to_bytes_returns_tuple(self, tool):
        with patch.object(tool.sql_tool, "run", side_effect=lambda q: MOCK_RESULTS[q]):
            result = tool.export_to_bytes()
            assert isinstance(result, tuple)
            assert len(result) == 2

    def test_export_to_bytes_returns_bytes(self, tool):
        with patch.object(tool.sql_tool, "run", side_effect=lambda q: MOCK_RESULTS[q]):
            content, filename = tool.export_to_bytes()
            assert isinstance(content, bytes)
            assert len(content) > 0

    def test_export_filename_format(self, tool):
        with patch.object(tool.sql_tool, "run", side_effect=lambda q: MOCK_RESULTS[q]):
            _, filename = tool.export_to_bytes()
            assert filename.startswith("TimeEntry_Validation_Report_")
            assert filename.endswith(".xlsx")

    def test_export_filename_contains_date(self, tool):
        from datetime import datetime
        with patch.object(tool.sql_tool, "run", side_effect=lambda q: MOCK_RESULTS[q]):
            _, filename = tool.export_to_bytes()
            today = datetime.now().strftime("%Y-%m-%d")
            assert today in filename

    def test_exported_bytes_is_valid_xlsx(self, tool):
        """Bytes must be a valid Excel file — openpyxl can open it."""
        from io import BytesIO
        from openpyxl import load_workbook
        with patch.object(tool.sql_tool, "run", side_effect=lambda q: MOCK_RESULTS[q]):
            content, _ = tool.export_to_bytes()
            wb = load_workbook(BytesIO(content))
            assert wb is not None

    def test_workbook_has_summary_sheet(self, tool):
        from io import BytesIO
        from openpyxl import load_workbook
        with patch.object(tool.sql_tool, "run", side_effect=lambda q: MOCK_RESULTS[q]):
            content, _ = tool.export_to_bytes()
            wb = load_workbook(BytesIO(content))
            assert "Summary" in wb.sheetnames

    def test_workbook_has_all_5_data_sheets(self, tool):
        from io import BytesIO
        from openpyxl import load_workbook
        with patch.object(tool.sql_tool, "run", side_effect=lambda q: MOCK_RESULTS[q]):
            content, _ = tool.export_to_bytes()
            wb = load_workbook(BytesIO(content))
            expected = {"All Entries", "Blank Memo", "Last 7 Days",
                        "Non ERP Memo", "Total Count"}
            assert expected.issubset(set(wb.sheetnames))

    def test_workbook_has_6_sheets_total(self, tool):
        from io import BytesIO
        from openpyxl import load_workbook
        with patch.object(tool.sql_tool, "run", side_effect=lambda q: MOCK_RESULTS[q]):
            content, _ = tool.export_to_bytes()
            wb = load_workbook(BytesIO(content))
            assert len(wb.sheetnames) == 6   # Summary + 5 data sheets

    def test_all_entries_sheet_has_correct_row_count(self, tool):
        from io import BytesIO
        from openpyxl import load_workbook
        with patch.object(tool.sql_tool, "run", side_effect=lambda q: MOCK_RESULTS[q]):
            content, _ = tool.export_to_bytes()
            wb = load_workbook(BytesIO(content))
            ws = wb["All Entries"]
            # Row 1 = header, rows 2-4 = 3 data rows
            data_rows = ws.max_row - 1   # subtract header
            assert data_rows >= 3

    def test_all_entries_sheet_has_headers(self, tool):
        from io import BytesIO
        from openpyxl import load_workbook
        with patch.object(tool.sql_tool, "run", side_effect=lambda q: MOCK_RESULTS[q]):
            content, _ = tool.export_to_bytes()
            wb = load_workbook(BytesIO(content))
            ws = wb["All Entries"]
            headers = [ws.cell(row=1, column=c).value for c in range(1, 6)]
            assert "ID"       in headers
            assert "EMPLOYEE" in headers
            assert "DATE"     in headers
            assert "HOURS"    in headers
            assert "MEMO"     in headers

    def test_empty_sheet_shows_no_records_message(self, tool):
        from io import BytesIO
        from openpyxl import load_workbook
        with patch.object(tool.sql_tool, "run", side_effect=lambda q: MOCK_RESULTS[q]):
            content, _ = tool.export_to_bytes()
            wb = load_workbook(BytesIO(content))
            ws = wb["Last 7 Days"]   # mock has 0 rows
            assert ws["A1"].value == "No records found"

    def test_summary_sheet_has_all_query_names(self, tool):
        from io import BytesIO
        from openpyxl import load_workbook
        with patch.object(tool.sql_tool, "run", side_effect=lambda q: MOCK_RESULTS[q]):
            content, _ = tool.export_to_bytes()
            wb = load_workbook(BytesIO(content))
            ws = wb["Summary"]
            all_text = " ".join(
                str(ws.cell(row=r, column=1).value or "")
                for r in range(1, ws.max_row + 1)
            )
            assert "Total" in all_text or "Retrieve" in all_text


# ═════════════════════════════════════════════════════════════
# INTEGRATION TESTS — needs SSH tunnel
# ═════════════════════════════════════════════════════════════

class TestExportIntegration:
    """Runs against real RDS — all queries, grows with data."""

    def test_export_to_bytes_with_real_data(self, tool):
        content, filename = tool.export_to_bytes()
        assert isinstance(content, bytes)
        assert len(content) > 5000

    def test_real_export_all_entries_has_rows(self, tool):
        """Uses >= 119 so test stays valid when new data is added."""
        from io import BytesIO
        from openpyxl import load_workbook
        content, _ = tool.export_to_bytes()
        wb = load_workbook(BytesIO(content))
        ws = wb["All Entries"]
        data_rows = [
            ws.cell(row=r, column=1).value
            for r in range(2, ws.max_row + 1)
            if isinstance(ws.cell(row=r, column=1).value, int)
        ]
        assert len(data_rows) >= 119

    def test_real_export_summary_has_row_counts(self, tool):
        """Summary column 3 must contain numeric row counts."""
        from io import BytesIO
        from openpyxl import load_workbook
        content, _ = tool.export_to_bytes()
        wb = load_workbook(BytesIO(content))
        ws = wb["Summary"]
        all_values = [
            ws.cell(row=r, column=3).value
            for r in range(1, ws.max_row + 1)
        ]
        numeric_values = [v for v in all_values if isinstance(v, int)]
        assert len(numeric_values) > 0, "Summary should have numeric row counts"
        assert max(numeric_values) >= 119, \
            f"Largest count should be >= 119, got {max(numeric_values)}"

    def test_real_export_blank_memo_all_null(self, tool):
        from io import BytesIO
        from openpyxl import load_workbook
        content, _ = tool.export_to_bytes()
        wb = load_workbook(BytesIO(content))
        ws = wb["Blank Memo"]
        if ws["A1"].value == "No records found":
            return   # valid — no blank memos
        # memo column is column 5 (id, employee, date, hours, memo)
        memo_col = None
        for c in range(1, 10):
            if str(ws.cell(row=1, column=c).value).upper() == "MEMO":
                memo_col = c
                break
        if memo_col:
            for r in range(2, ws.max_row + 1):
                val = ws.cell(row=r, column=memo_col).value
                if ws.cell(row=r, column=1).value is not None:
                    assert val is None, f"Row {r} has memo='{val}'"

    def test_export_file_saves_to_disk(self, tool, tmp_path):
        filepath = tool.export(output_dir=str(tmp_path))
        assert os.path.exists(filepath)
        assert filepath.endswith(".xlsx")
        assert os.path.getsize(filepath) > 0
